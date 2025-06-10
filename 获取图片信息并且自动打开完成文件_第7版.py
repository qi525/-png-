# -*- coding: utf-8 -*-
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re
from PIL import Image, ImageFile
from datetime import datetime
import subprocess
import warnings # 导入warnings模块
import inspect # 导入inspect模块用于检查调用栈

# [2025-06-10] 永远不要使用等待加载"networkidle"。
# [2025-06-10] 原有的功能不要乱改，应该多增加代码，少去删除以前的代码，没说让你改的地方别乱改，优先处理我说的问题。
# [2025-06-08] 代码一定是方便调试的，报错给我生成txt的log。存储信息用xlsx。log的txt和xlsx完成代码的时候要自动运行打开方便检查结果。
# [2025-06-08] 我提供的代码不要删除我的注释

# 允许 Pillow 加载截断的图像文件，避免程序崩溃。
# 注意：即使设置为 True，如果文件严重损坏，仍可能引发异常，被下面的try-except捕获。
# 且UserWarning本身不会被try-except捕获，需要额外的warnings处理。
ImageFile.LOAD_TRUNCATED_IMAGES = True

# 全局变量，用于在警告处理函数中访问当前处理的文件路径
_current_processing_file = None

def custom_warning_formatter(message, category, filename, lineno, file=None, line=None):
    """
    自定义警告格式化器，尝试获取当前处理的文件路径。
    """
    global _current_processing_file
    
    # 检查警告是否来自 PIL 的 TiffImagePlugin 并且是 Truncated File Read
    if category is UserWarning and "Truncated File Read" in str(message) and "TiffImagePlugin.py" in filename:
        if _current_processing_file:
            return f"UserWarning: {message} for file: '{_current_processing_file}'\n"
    
    # 对于其他警告，使用默认格式
    return warnings.formatwarning(message, category, filename, lineno, line)

# 设置自定义警告格式化器
warnings.formatwarning = custom_warning_formatter

def log_error(message):
    """
    记录错误信息到控制台和日志文件。
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"{timestamp} - {message}\n"
    print(message) # 在控制台打印错误信息
    with open("image_scan_error.log", "a", encoding="utf-8") as log_file:
        log_file.write(log_entry)

def get_image_info(folder_path):
    """
    Scans a folder for image files, extracts their paths, parent folders (absolute path),
    and Stable Diffusion generation information.
    """
    global _current_processing_file # 声明使用全局变量

    image_data = []
    image_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')

    # 定义一个更通用的正则表达式，用于从原始文本中捕获 Stable Diffusion 的信息块
    # 它会从常见的提示词或Negative prompt开始匹配，直到最后一个参数Version结束
    sd_full_info_pattern = re.compile(
        r'.*?(?:masterpiece|score_\d|1girl|BREAK|Negative prompt:|Steps:).*?(?:Version:.*?|Module:.*?|)$',
        re.DOTALL # 允许.匹配换行符
    )
    # 定义一个更严格的正则，用于最终验证是否是有效的SD参数
    sd_validation_pattern = re.compile(r'Steps: \d+, Sampler: [\w\s]+', re.DOTALL)


    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(image_extensions):
                absolute_path = os.path.abspath(os.path.join(root, file))
                containing_folder_absolute_path = os.path.abspath(root)
                sd_info = "没有扫描到生成信息" # 默认值
                positive_prompt = ""
                negative_prompt = ""
                other_settings = ""

                raw_metadata_string = "" # 用于存储从图片中初步提取的原始字符串

                _current_processing_file = absolute_path # 在处理每个文件前更新全局变量

                try:
                    # 尝试打开图像文件。如果文件损坏或截断，Image.open()可能会引发IOError或类似的异常
                    with Image.open(absolute_path) as img:
                        # --- 阶段 1: 尝试从标准位置获取原始元数据字符串 ---
                        if "png" in img.format.lower() and "parameters" in img.info:
                            raw_metadata_string = img.info["parameters"]
                        elif "jpeg" in img.format.lower():
                            if hasattr(img, '_getexif'):
                                exif_data = img._getexif()
                                if exif_data:
                                    for tag, value in exif_data.items():
                                        if tag in [0x9286, 0x010E]: # UserComment or ImageDescription
                                            try:
                                                # 尝试UTF-8解码，这是最常见的编码
                                                raw_metadata_string = value.decode('utf-8', errors='ignore')
                                                # 如果解码后仍然没有明显的SD参数特征，可以尝试其他编码
                                                if not re.search(r'Steps:', raw_metadata_string):
                                                    raw_metadata_string = value.decode('latin-1', errors='ignore')
                                                break # 找到就跳出
                                            except Exception:
                                                pass
                        
                        # --- 阶段 2: 清理并使用更强大的正则表达式提取有效信息 ---
                        if isinstance(raw_metadata_string, str) and raw_metadata_string:
                            # 移除 Excel 不支持的非法 XML 字符
                            cleaned_string = ILLEGAL_CHARACTERS_RE.sub(r'', raw_metadata_string)
                            
                            # Clean up the "UNICODE" prefix
                            if cleaned_string.startswith("UNICODE"):
                                cleaned_string = cleaned_string[len("UNICODE"):].lstrip() # Remove "UNICODE" and any leading whitespace
                            
                            # 尝试使用新的正则表达式捕获核心SD信息块
                            match = sd_full_info_pattern.search(cleaned_string)
                            
                            if match:
                                extracted_text = match.group(0).strip() # 获取匹配到的整个SD信息块
                                # 再次使用更严格的正则验证，确保提取的是有效的SD参数
                                if sd_validation_pattern.search(extracted_text):
                                    sd_info = extracted_text
                                    
                                    # --- 阶段 3: 切割信息 ---
                                    # 从后往前切割
                                    other_settings_match = re.search(r'(Steps:.*)', sd_info, re.DOTALL)
                                    if other_settings_match:
                                        other_settings = other_settings_match.group(1).strip()
                                        temp_sd_info = sd_info[:other_settings_match.start()].strip()
                                    else:
                                        temp_sd_info = sd_info.strip() # 如果没有Steps，则整个认为是正向提示词

                                    negative_prompt_match = re.search(r'(Negative prompt:.*?)(?=\nSteps:|$)', temp_sd_info, re.DOTALL)
                                    if negative_prompt_match:
                                        negative_prompt = negative_prompt_match.group(1).replace("Negative prompt:", "").strip()
                                        positive_prompt = temp_sd_info[:negative_prompt_match.start()].strip()
                                    else:
                                        positive_prompt = temp_sd_info.strip() # 如果没有Negative prompt，则整个认为是正向提示词

                                else:
                                    # 即使匹配到了，但最终验证不通过，也认为没有扫描到
                                    sd_info = "没有扫描到生成信息"
                            else:
                                # 如果通用模式都无法匹配到，那就不包含SD信息
                                sd_info = "没有扫描到生成信息"

                except Exception as e:
                    # 如果Image.open()或后续操作因文件损坏而失败，这里的e会包含详细错误信息
                    log_error(f"Error processing image file '{absolute_path}': {e}") # 明确指出是哪个文件出了问题
                    sd_info = "没有扫描到生成信息" # 发生任何错误时都重置
                    positive_prompt = ""
                    negative_prompt = ""
                    other_settings = ""
                finally:
                    _current_processing_file = None # 处理完一个文件后重置全局变量

                image_data.append({
                    "所在文件夹": containing_folder_absolute_path,
                    "图片的绝对路径": absolute_path,
                    "图片超链接": f'={absolute_path}',
                    "stable diffusion的 ai图片的生成信息": sd_info,
                    "正面提示词": positive_prompt,
                    "负面提示词": negative_prompt,
                    "其他设置": other_settings
                })
    return image_data

def create_excel_report(image_data, base_filename="图片信息报告"):
    """
    Creates an Excel report from the collected image data with a timestamped filename
    and attempts to open it automatically.
    """
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_filename = f"{base_filename}_{timestamp}.xlsx"

    df = pd.DataFrame(image_data)

    if df.empty:
        print("没有找到任何图片文件，将创建一个空的Excel文件。")
        df = pd.DataFrame(columns=[
            "所在文件夹",
            "图片的绝对路径",
            "图片超链接",
            "stable diffusion的 ai图片的生成信息",
            "正面提示词",
            "负面提示词",
            "其他设置"
        ])

    writer = pd.ExcelWriter(output_filename, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='图片信息')

    workbook = writer.book
    sheet = writer.sheets['图片信息']

    for col_idx, column_name in enumerate(df.columns):
        max_length = 0
        # Set a default minimum width for all columns
        adjusted_width = 15
        
        # Calculate max_length for content, excluding "图片超链接" for length calculation
        if column_name != "图片超链接":
            for cell_value in df[column_name]:
                try:
                    length = len(str(cell_value))
                    if length > max_length:
                        max_length = length
                except TypeError:
                    pass
            # Adjust width based on content or header length
            adjusted_width = max(max_length + 2, len(column_name) + 2)
        else: # For "图片超链接", base width on the original path length
            for row_idx in range(len(df["图片的绝对路径"])):
                try:
                    length = len(str(df["图片的绝对路径"].iloc[row_idx]))
                    if length > max_length:
                        max_length = length
                except IndexError:
                    pass
            adjusted_width = max(max_length + 2, len(column_name) + 2) # Also consider header length


        sheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width

        if column_name == "图片超链接":
            for row_idx, cell_value in enumerate(df[column_name]):
                cell = sheet.cell(row=row_idx + 2, column=col_idx + 1)
                original_path = df["图片的绝对路径"].iloc[row_idx]
                cell.hyperlink = f"file:///{original_path}"
                cell.value = "点击查看原图"
                cell.font = Font(color=Color("0000FF"), underline="single")

    writer.close()
    print(f"数据已成功保存到 {output_filename}")

    try:
        if os.name == 'nt':  # For Windows
            os.startfile(output_filename)
        elif os.uname().sysname == 'Darwin':  # For macOS
            subprocess.run(['open', output_filename], check=True)
        else:  # For Linux (assuming xdg-open is available)
            subprocess.run(['xdg-open', output_filename], check=True)
        print(f"尝试自动打开文件: {output_filename}")
    except FileNotFoundError:
        print(f"错误: 无法找到打开 '{output_filename}' 的应用程序。请手动打开。")
    except Exception as e:
        print(f"自动打开文件时发生错误: {e}")

if __name__ == "__main__":
    folder_to_scan = input("请输入要扫描的文件夹路径: ")

    if not os.path.isdir(folder_to_scan):
        print(f"错误: 文件夹 '{folder_to_scan}' 不存在。请提供一个有效的文件夹路径。")
        log_error(f"用户输入的文件夹 '{folder_to_scan}' 不存在。") # 记录文件夹不存在的错误
    else:
        print(f"正在扫描文件夹: {folder_to_scan}...")
        image_info = get_image_info(folder_to_scan)
        create_excel_report(image_info)