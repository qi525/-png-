import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re
from PIL import Image

def get_image_info(folder_path):
    """
    Scans a folder for image files, extracts their paths, parent folders,
    and Stable Diffusion generation information.
    """
    image_data = []
    image_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp') # Add more if needed

    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(image_extensions):
                absolute_path = os.path.abspath(os.path.join(root, file))
                containing_folder = os.path.basename(root)
                sd_info = "没有扫描到生成信息"

                try:
                    with Image.open(absolute_path) as img:
                        # Attempt to get EXIF data or other metadata for Stable Diffusion info
                        # Stable Diffusion often embeds info in the 'parameters' or 'Description' field of PNGs,
                        # or in EXIF UserComment for JPGs.
                        if "png" in img.format.lower() and "parameters" in img.info:
                            sd_info = img.info["parameters"]
                        elif "jpeg" in img.format.lower():
                            if hasattr(img, '_getexif'):
                                exif_data = img._getexif()
                                if exif_data:
                                    for tag, value in exif_data.items():
                                        if tag in [0x9286, 0x010E]: # UserComment or ImageDescription
                                            try:
                                                decoded_value = value.decode('utf-8')
                                                if "parameters" in decoded_value or "Steps:" in decoded_value:
                                                    sd_info = decoded_value
                                                    break
                                            except (UnicodeDecodeError, AttributeError):
                                                pass
                        # Clean up potentially illegal XML characters from the string
                        if isinstance(sd_info, str):
                            sd_info = ILLEGAL_CHARACTERS_RE.sub(r'', sd_info)
                            # Basic regex to find common Stable Diffusion parameters
                            if not re.search(r'Steps: \d+, Sampler: \w+', sd_info):
                                # If typical SD info isn't found, revert to default message
                                sd_info = "没有扫描到生成信息"

                except Exception as e:
                    print(f"Error processing {absolute_path}: {e}")
                    sd_info = "没有扫描到生成信息"

                image_data.append({
                    "所在文件夹": containing_folder,
                    "图片的绝对路径": absolute_path,
                    "图片超链接": f'={absolute_path}', # This will be adjusted to a proper hyperlink in Excel
                    "stable diffusion的 ai图片的生成信息": sd_info
                })
    return image_data

def create_excel_report(image_data, output_filename="图片信息报告.xlsx"):
    """
    Creates an Excel report from image data.
    """
    df = pd.DataFrame(image_data)

    if df.empty:
        print("没有找到任何图片文件，将创建一个空的Excel文件。")
        df = pd.DataFrame(columns=[
            "所在文件夹",
            "图片的绝对路径",
            "图片超链接",
            "stable diffusion的 ai图片的生成信息"
        ])


    writer = pd.ExcelWriter(output_filename, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='图片信息')

    workbook = writer.book
    sheet = writer.sheets['图片信息']

    # Set column widths and add hyperlinks
    for col_idx, column_name in enumerate(df.columns):
        max_length = 0
        for row_idx, cell_value in enumerate(df[column_name]):
            # Adjust max_length calculation for hyperlinks
            if column_name == "图片超链接":
                # For hyperlinks, the displayed text might be different from the actual path
                # Let's base width on the absolute path column for better visibility
                if row_idx < len(df["图片的绝对路径"]):
                    try:
                        length = len(str(df["图片的绝对路径"].iloc[row_idx]))
                        if length > max_length:
                            max_length = length
                    except IndexError:
                        pass
            else:
                try:
                    length = len(str(cell_value))
                    if length > max_length:
                        max_length = length
                except TypeError: # Handle cases where cell_value might not be a string
                    pass
        adjusted_width = (max_length + 2) if max_length > 0 else 15 # Add a little padding
        sheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width

        # Apply hyperlink formatting to column C
        if column_name == "图片超链接":
            for row_idx, cell_value in enumerate(df[column_name]):
                cell = sheet.cell(row=row_idx + 2, column=col_idx + 1) # +2 for header row and 0-indexed col
                original_path = df["图片的绝对路径"].iloc[row_idx]
                cell.hyperlink = f"file:///{original_path}"
                cell.value = "点击查看原图"
                cell.font = Font(color=Color("0000FF"), underline="single")


    writer.close()
    print(f"数据已成功保存到 {output_filename}")

if __name__ == "__main__":
    folder_to_scan = input("请输入要扫描的文件夹路径: ")

    if not os.path.isdir(folder_to_scan):
        print(f"错误: 文件夹 '{folder_to_scan}' 不存在。")
    else:
        print(f"正在扫描文件夹: {folder_to_scan}...")
        image_info = get_image_info(folder_to_scan)
        create_excel_report(image_info)