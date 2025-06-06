import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re
from PIL import Image
from datetime import datetime
import subprocess # Import subprocess for cross-platform opening

def get_image_info(folder_path):
    """
    Scans a folder for image files, extracts their paths, parent folders (absolute path),
    and Stable Diffusion generation information.
    """
    image_data = []
    image_extensions = ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')

    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith(image_extensions):
                absolute_path = os.path.abspath(os.path.join(root, file))
                containing_folder_absolute_path = os.path.abspath(root)
                sd_info = "没有扫描到生成信息"

                try:
                    with Image.open(absolute_path) as img:
                        if "png" in img.format.lower() and "parameters" in img.info:
                            sd_info = img.info["parameters"]
                        elif "jpeg" in img.format.lower():
                            if hasattr(img, '_getexif'):
                                exif_data = img._getexif()
                                if exif_data:
                                    for tag, value in exif_data.items():
                                        if tag in [0x9286, 0x010E]:
                                            try:
                                                decoded_value = value.decode('utf-8')
                                                if "parameters" in decoded_value or "Steps:" in decoded_value:
                                                    sd_info = decoded_value
                                                    break
                                            except (UnicodeDecodeError, AttributeError):
                                                pass
                        
                        if isinstance(sd_info, str):
                            sd_info = ILLEGAL_CHARACTERS_RE.sub(r'', sd_info)
                            if not re.search(r'Steps: \d+, Sampler: \w+', sd_info):
                                sd_info = "没有扫描到生成信息"

                except Exception as e:
                    print(f"Error processing {absolute_path}: {e}")
                    sd_info = "没有扫描到生成信息"

                image_data.append({
                    "所在文件夹": containing_folder_absolute_path,
                    "图片的绝对路径": absolute_path,
                    "图片超链接": f'={absolute_path}',
                    "stable diffusion的 ai图片的生成信息": sd_info
                })
    return image_data

def create_excel_report(image_data, base_filename="图片信息报告"):
    """
    Creates an Excel report from the collected image data with a timestamped filename
    and attempts to open it automatically.
    """
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_filename = f"{base_filename}_{timestamp}.xlsx"

    df = pd.DataFrame(image_data)

    if df.empty:
        print("没有找到任何图片文件，将创建一个空的Excel文件。")
        df = pd.DataFrame(columns=[
            "所在文件夹",
            "图片的绝对路径",
            "图片超链接",
            "stable diffusion的 ai图片的生成信息"
        ])

    writer = pd.ExcelWriter(output_filename, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='图片信息')

    workbook = writer.book
    sheet = writer.sheets['图片信息']

    for col_idx, column_name in enumerate(df.columns):
        max_length = 0
        for row_idx, cell_value in enumerate(df[column_name]):
            if column_name == "图片超链接":
                if row_idx < len(df["图片的绝对路径"]):
                    try:
                        length = len(str(df["图片的绝对路径"].iloc[row_idx]))
                        if length > max_length:
                            max_length = length
                    except IndexError:
                        pass
            else:
                try:
                    length = len(str(cell_value))
                    if length > max_length:
                        max_length = length
                except TypeError:
                    pass
        adjusted_width = (max_length + 2) if max_length > 0 else 15
        sheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width

        if column_name == "图片超链接":
            for row_idx, cell_value in enumerate(df[column_name]):
                cell = sheet.cell(row=row_idx + 2, column=col_idx + 1)
                original_path = df["图片的绝对路径"].iloc[row_idx]
                cell.hyperlink = f"file:///{original_path}"
                cell.value = "点击查看原图"
                cell.font = Font(color=Color("0000FF"), underline="single")

    writer.close()
    print(f"数据已成功保存到 {output_filename}")

    # --- New Functionality: Auto-Open the XLSX file ---
    try:
        if os.name == 'nt':  # For Windows
            os.startfile(output_filename)
        elif os.uname().sysname == 'Darwin':  # For macOS
            subprocess.run(['open', output_filename], check=True)
        else:  # For Linux (assuming xdg-open is available)
            subprocess.run(['xdg-open', output_filename], check=True)
        print(f"尝试自动打开文件: {output_filename}")
    except FileNotFoundError:
        print(f"错误: 无法找到打开 '{output_filename}' 的应用程序。请手动打开。")
    except Exception as e:
        print(f"自动打开文件时发生错误: {e}")
    # --- End New Functionality ---

if __name__ == "__main__":
    folder_to_scan = input("请输入要扫描的文件夹路径: ")

    if not os.path.isdir(folder_to_scan):
        print(f"错误: 文件夹 '{folder_to_scan}' 不存在。请提供一个有效的文件夹路径。")
    else:
        print(f"正在扫描文件夹: {folder_to_scan}...")
        image_info = get_image_info(folder_to_scan)
        create_excel_report(image_info)